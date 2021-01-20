from flask import Flask,request
import openpyxl
import datetime
import os.path
app = Flask(__name__)
pwd = os.path.abspath(os.path.dirname(__file__))
max_buf=1
ticha=[]
@app.route('/', methods=['POST', 'GET'])
def index():
	if request.method == 'POST':
		user_id=request.json['user_id']
		check=request.json['check']
		for a in range(len(check)):
			ticha.append(check[a])
			ticha[len(ticha)-1]["user_id"]=user_id
			ticha[len(ticha)-1]["datetime"]= datetime.datetime.now()
		if(len(ticha)>=max_buf):

			STORAGE_FILE = os.path.join(pwd, 'data.xlsx')
			book = None
			
			if not os.path.exists(STORAGE_FILE):
				book = openpyxl.Workbook()
				book.active['A1'] = 'N'
				book.active['B1'] = 'User ID'
				book.active['C1'] = 'Datetime'
				book.active['D1'] = 'Item'
				book.active['E1'] = 'Prise'
				book.save(STORAGE_FILE)
			else:
				book = openpyxl.open(STORAGE_FILE)   
			sheet = book.active
			rows=len(sheet['A'])
			start=rows+1
			for i in range (0,len(ticha)):
				sheet.cell(row=i+start,column=1).value = i+1+start
				sheet.cell(row=i+start,column=2).value = ticha[i]["user_id"]
				sheet.cell(row=i+start,column=3).value = ticha[i]["datetime"]
				sheet.cell(row=i+start,column=4).value = ticha[i]["item"]
				sheet.cell(row=i+start,column=5).value = ticha[i]["price"]		
			book.save("data.xlsx")
			book.close()
			ticha.clear()
		return "OK"
	if request.method == 'GET':
		return "Это GET запрос"
 
if __name__ == "__main__":
	app.run()